import concurrent.futures
import copy
import random
import socket
import struct
import sys
import time
import ipaddress
import threading
import warnings

# Filter out the specific warning message
warnings.filterwarnings("ignore", message="Your system is avx2 capable but pygame was not built with support for it.")

import pygame

from colorama import Fore

from Bot import Bot

from openpyxl import load_workbook

"""
This script implements a server for the Trivia King game.

The server communicates with clients over TCP/IP and broadcasts offers via UDP messages.
Players can join the game by responding to the UDP broadcast and connecting to the server.
The game consists of answering trivia questions about geography.

The script includes the following functionalities:
1. UDP Broadcast: The server broadcasts offers to potential clients using UDP messages.
2. TCP Server: Manages client connections, handles game logic, and sends trivia questions to clients.
3. Bot Support: Allows adding bots to the game in case there are not enough human players.

The server uses the following external libraries:
- colorama: for colored console output
- pygame: for playing sounds
- openpyxl: for updating and retrieving data from an Excel file to maintain player scores

The server functions include:
- udp_broadcast: Broadcasts server offers via UDP messages.
- tcp_server: Manages client connections, handles game logic, and communicates with clients over TCP/IP.
- play_sound: Plays a sound file.
- update_excel: Updates an Excel file with player scores.
- get_top_three_players: Retrieves the top three players from the Excel file.
- pad_server_name: Pads the server name to a fixed length for broadcasting.
- get_local_ipv4_address: Retrieves the local IPv4 address of the server.
- find_available_port: Finds an available port for the server to bind to.
- get_local_broadcast_ip: Retrieves the local broadcast IP address.
- check_if_disconnected: Checks if any clients have disconnected.
- everyone_wrong_or_right: Checks if all players either answered correctly or incorrectly.
- safe_sendall: Safely sends a message to all clients, handling potential socket errors.
- send_to_all: Sends a message to all clients.
- elimination_msg: Sends an elimination message to a specific client.
- send_question: Sends a trivia question to a specific client.
- play_trivia: Manages the trivia game for a specific client.
"""


def play_sound(sound_file):
    """
    Play a sound file.
    :param sound_file:
    :return:
    """
    pygame.mixer.init()
    sound = pygame.mixer.Sound(sound_file)
    sound.play()

def update_excel(filename, name):
    """
    Update the Excel file by incrementing the wins count for the specified name.

    Args:
    - filename: The filename of the Excel file.
    - name: The name of the winner.
    """
    try:
        # Load the workbook with write access
        wb = load_workbook(filename, read_only=False)
        ws = wb.active

        # Check if the name already exists
        name_exists = False

        for row in ws.iter_rows(min_row=1, max_col=1, max_row=ws.max_row, values_only=True):
            if row[0] == name:
                name_exists = True

                break

        if name_exists:
            i = 1
            for row in ws.iter_rows(values_only=True):
                if row[0] == name:
                    new_score = row[1] + 1
                    ws.cell(row=i, column=2, value=new_score)
                    break
                i += 1

            # Save the changes to the Excel file
            wb.save(filename)
        else:
            # Add a new row for the name
            ws.append([name, 1])

        # Save the workbook
        wb.save(filename)

    except Exception as e:
        pass


def get_top_three_players(filename):
    """
    Retrieve the top three players with the most wins from the Excel file.

    Args:
    - filename: The filename of the Excel file.

    Returns:
    A list of tuples containing the top three players with their names and wins.
    """
    try:
        # Load the workbook
        wb = load_workbook(filename, read_only=True)
        ws = wb.active

        # Create a dictionary to store player names and wins
        players = {}

        # Read data from the Excel file
        for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row, values_only=True):
            name, wins = row
            if name in players:
                players[name] += wins
            else:
                players[name] = wins

        # Sort players by wins in descending order
        sorted_players = sorted(players.items(), key=lambda x: x[1], reverse=True)

        # Get the top three players
        top_three = sorted_players[:3]

        return top_three

    except Exception as e:
        pass


# Function to pad server name to 32 characters
def pad_server_name(server_name):
    return server_name.ljust(32, '\0')


def get_local_ipv4_address():
    """
    Get the local IPv4 address of the server.
    :return:
    """
    # Get the local hostname
    temp_sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    temp_sock.connect(("8.8.8.8", 80))
    return temp_sock.getsockname()[0]

def find_available_port():
    """
    Find an available port for the server to bind to.
    :return:
    """
    # Create a socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        # Bind the socket to a random port
        s.bind(('', 0))
        return s.getsockname()[1]


# Define constants
MAGIC_COOKIE = 0xabcddcba
MESSAGE_TYPE_OFFER = 0x2
LOCAL_IP = get_local_ipv4_address()
server_name = pad_server_name("Lucky Bunnies")
available_port = find_available_port()

def get_local_broadcast_ip():
    """
    Get the local broadcast IP address.
    :return: broadcast address of the network
    """
    # Get the local IP address
    ip = LOCAL_IP
    # Get the local network mask
    mask = ipaddress.IPv4Network(ip + '/24', strict=False)
    # Get the broadcast IP address
    broadcast_ip = str(mask.broadcast_address)
    return broadcast_ip


def udp_broadcast():
    """
        Broadcasts server offers via UDP messages.

        The function broadcasts offers containing the server name and port to potential clients.
        It uses a UDP socket to send messages to the local broadcast IP address.

        """
    BROADCAST_IP = get_local_broadcast_ip()
    BROADCAST_PORT = 13117
    server_port = available_port
    packed_data = struct.pack('!IB32sH', MAGIC_COOKIE, MESSAGE_TYPE_OFFER, server_name.encode('utf-8'), server_port)

    # Create a UDP socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    # Set the socket to broadcast and enable reusing addresses
    sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
    try:
        print(f'Server started, listening on IP address: {LOCAL_IP}')
        time.sleep(0.5)
        while True:
            # Send new message every second
            sock.sendto(packed_data, (BROADCAST_IP, BROADCAST_PORT))
            time.sleep(1)
    except KeyboardInterrupt:
        print('Stopping broadcast')
        # Close the socket
        sock.close()
    except OSError:
        sock.close()


questions = {
    "The capital city of Japan is Tokyo": True,
    "The Amazon River is the longest river in the world": False,
    "The Great Wall of China is the longest wall in the world": True,
    "The Sahara Desert is the largest hot desert in the world": True,
    "The tallest mountain in the world is Mount Everest": True,
    "The largest country in the world by land area is Russia": True,
    "The city of Rome is located in France": False,
    "The Nile River is the longest river in the world": False,
    "The Eiffel Tower is located in London": False,
    "The Dead Sea is the saltiest body of water in the world": True,
    "The Statue of Liberty was a gift from France to the United States": True,
    "The Sydney Opera House is located in Australia": True,
    "The capital city of Canada is Ottawa": True,
    "The Sahara Desert is located in South America": False,
    "The largest ocean in the world is the Pacific Ocean": True,
    "The city of Moscow is the capital of Russia": True,
    "The Great Barrier Reef is located in the Indian Ocean": False,
    "The city of Berlin is the capital of Germany": True,
    "The Panama Canal connects the Pacific Ocean to the Atlantic Ocean": True,
    "The highest waterfall in the world is Angel Falls": True,
}

bot_names = [
    "BOT_columbus",
    "BOT_magellan",
    "BOT_cook",
    "BOT_vespucci",
    "BOT_hudson",
    "BOT_cabot",
    "BOT_drake",
    "BOT_marco_polo",
    "BOT_champlain",
    "BOT_cortes",
    "BOT_pizarro",
    "BOT_cartier",
    "BOT_la_salle",
    "BOT_park",
    "BOT_livingstone",
    "BOT_vasco_da_gama",
    "BOT_pedro_alvares_cabral",
    "BOT_zheng_he",
    "BOT_meriwether_lewis",
    "BOT_william_clark"
]

acceptable_answers = {'0': 'F', '1': 'T', 'Y': 'T', 'N': 'F', 'T': 'T', 'F': 'F', 'NONE': None}
disconnected_clients = {}
pid_to_name = {}


def tcp_server(host, port, fill_bots=True, add_bots=0):
    """
     Manages client connections, handles game logic, and communicates with clients over TCP/IP.

     The function listens for incoming client connections and creates a new thread for each client.
     It handles game logic, including sending trivia questions, receiving answers, and managing player eliminations.
     Bots can be added to the game automatically if there are not enough human players.

     Args:
     - host (str): The IP address of the server.
     - port (int): The port number to bind the server socket.
     - fill_bots (bool): Whether to add bots to the game if there are not enough human players (default=True).
     - add_bots (int): The number of additional bots to add to the game (default=0).

     Returns:
     None
     """
    global pid_to_name
    global disconnected_clients
    welcome_msg = "Please wait for other players to join...\n"

    # Function to send welcome message to all the clients
    def handle_client(client_id, client_socket, players):
        """
        Handles client connections and game logic for each client.
        :param client_id: id of the client joining
        :param client_socket: socket of the client joining
        :param players: dict of players
        :return:
        """
        try:
            player_name = client_socket.recv(1024)
            client_socket.sendall(welcome_msg.encode())
            print(f"Player {player_name.decode()} joined the lobby.\n")
            send_to_all(client_sockets_og, f"Player {player_name.decode()} joined the lobby.\n")
            players[client_id] = (player_name.decode())
        except OSError:
            global disconnected_clients
            try:
                disconnected_clients[client_id] = players[client_id]
            except Exception:
                pass

    # Create a TCP/IP socket
    server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server_socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)

    # Bind the socket to the address and port

    con = False
    while not con:
        try:
            server_socket.bind((host, port))
            con = True
        except OSError:

            server_socket.close()
            server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)

    try:
        while True:
            # Listen for incoming connections
            server_socket.listen()
            pid_to_name = {}
            players = {}
            client_sockets_og = {}
            # Accept incoming connections, create a new thread for each client,
            # and welcome them, while waiting for all players to join the lobby.
            i = 0
            while True:
                client_handlers = []
                try:
                    i += 1
                    client_socket, client_address = server_socket.accept()
                    client_handler = threading.Thread(target=handle_client, args=(i, client_socket, players))
                    client_sockets_og[i] = client_socket
                    client_handler.start()
                    client_handlers.append(client_handler)
                    server_socket.settimeout(10)
                except socket.timeout:
                    # If no new connections are made within the 10 second period, start the game
                    server_socket.settimeout(None)
                    for t in client_handlers:
                        t.join()
                    for client_id, sock in client_sockets_og.copy().items():
                        try:
                            # Check if the socket is still connected
                            sock.sendall("".encode())
                        except OSError:
                            # If an OSError is raised, the socket is no longer connected
                            print(f"Player {players[client_id]} is no longer connected")
                            client_sockets_og.pop(client_id)
                            players.pop(client_id)
                    break

            client_sockets = dict(client_sockets_og)
            time.sleep(1)
            check_if_disconnected(client_sockets, client_sockets_og, players)

            if len(client_sockets) == 0:
                print('No players connected, waiting for new players...\n\n')
                continue
            elif len(client_sockets) < 4 and fill_bots:
                print('Not enough players. Adding bots to the game.')
                send_to_all(client_sockets, 'Not enough players. Adding bots to the game.')
                i = len(client_sockets)
                for j in range(4 - len(client_sockets)):
                    i += 1
                    add_bots -= 1
                    server_socket.listen()
                    name = random.choice(bot_names)
                    bot_thread = threading.Thread(target=Bot(name, address=LOCAL_IP, server_port=available_port, isBot=True).run)
                    bot_thread.start()
                    client_socket, client_address = server_socket.accept()
                    client_handler = threading.Thread(target=handle_client, args=(i, client_socket, players))
                    client_sockets_og[i] = client_socket
                    client_sockets[i] = client_socket
                    client_handler.start()
                    client_handler.join()

            if add_bots > 0:
                i = len(client_sockets)
                for j in range(add_bots):
                    i += 1
                    server_socket.listen()
                    name = random.choice(bot_names)
                    bot_thread = threading.Thread(target=Bot(name, address=LOCAL_IP, server_port=available_port, isBot=True).run)
                    bot_thread.start()
                    client_socket, client_address = server_socket.accept()
                    client_handler = threading.Thread(target=handle_client, args=(i, client_socket, players))
                    client_sockets_og[i] = client_socket
                    client_sockets[i] = client_socket
                    client_handler.start()
                    client_handler.join()

            time.sleep(1)
            check_if_disconnected(client_sockets, client_sockets_og, players)

            print('Welcome to LuckyBunnies Server, where we are answering trivia questions about geography!\n')
            send_to_all(client_sockets,'Welcome to LuckyBunnies Server, where we are answering trivia questions about geography!\n')
            print("Loading game...\n")
            send_to_all(client_sockets,"Loading game...")
            time.sleep(1)
            message = ""
            pid_to_name = dict(players)

            for i, key in enumerate(pid_to_name.keys()):
                message += f"Player {i + 1}: {pid_to_name[key]}\n"
            message += "===============\n"
            print(message)
            curr_threads = []
            # send the players list to all the clients
            send_to_all(client_sockets, message)

            curr_threads = []
            time.sleep(2)
            # Start the game: send questions to all the clients and wait for their answers
            i = 1
            for question in questions:
                # check if any clients have disconnected
                for dc in disconnected_clients:
                    print(f'{disconnected_clients[dc]} has disconnected from the game.')
                    send_to_all(client_sockets,
                                f'{disconnected_clients[dc]} has disconnected from the game.\n')
                    pid_to_name.pop(dc)
                    client_sockets.pop(dc)

                disconnected_clients = {}

                check_if_disconnected(client_sockets, client_sockets_og, pid_to_name)

                if len(pid_to_name) < 2:
                    break

                next_round = f'Round {i}, played by '
                next_round += ', '.join([pid_to_name[player] for player in client_sockets.keys()]) + ':'
                # print(next_round)
                # send_to_all(client_sockets, f'{next_round}\n')
                i += 1
                res = play(client_sockets, pid_to_name, question, curr_threads, next_round=next_round)
                if res is True:
                    if len(pid_to_name) == 1:
                        break

            if len(pid_to_name) == 0:
                send_to_all(client_sockets_og, "Game is tied !")
                print("No players left in the game.\nGame over, sending out offer requests...\n\n")
                time.sleep(1)

            elif len(pid_to_name) > 1:
                print("Game is tied !\nLooking for new players... ")
                send_to_all(client_sockets_og, "Game is tied !")
                time.sleep(1)


            elif len(pid_to_name) == 1:
                name = next(iter(pid_to_name.values()))
                time.sleep(2)
                update_excel('winners.xlsx', name)
                top_three_players = get_top_three_players('winners.xlsx')
                game_over_mess = f'Game over!\nCongratulations to the winner: {name}\nAll Time Server Rankings:\n'
                i = 1
                for player in top_three_players:
                    game_over_mess += f'{i}. {player[0]}: {player[1]}\n'
                    i += 1
                print(game_over_mess)
                send_to_all(client_sockets_og, game_over_mess)
                time.sleep(2)
                print("Game over, sending out offer requests...\n\n")
                time.sleep(1)

    except KeyboardInterrupt:
        server_socket.close()
        print("Shutting down the server... Goodbye!")


def play(client_sockets, pid_to_name, question, curr_threads, next_round=""):
    """
        Manages a round of the trivia game.

        The function sends the trivia question to all clients, waits for their answers,
        determines the correctness of each answer, and handles player eliminations.

        Args:
        - client_sockets (dict): A dictionary containing client IDs and their corresponding sockets.
        - pid_to_name (dict): A dictionary mapping player IDs to their names.
        - question (str): The trivia question to be asked.
        - curr_threads (list): A list of current active threads.
        - next_round (str): Information about the next round (default="").

        Returns:
        True if no players are eliminated in the round.
        """

    print(next_round)
    send_to_all(client_sockets, f'{next_round}\n')
    time.sleep(5)
    print(f"True or False: {question}\n")
    client_answers = {}
    # for client_id, sock in client_sockets.items():
    #     t = threading.Thread(target=send_question, args=(client_id, sock, question))
    #     t.start()
    #     curr_threads.append(t)
    # for t in curr_threads:
    #     t.join()
    mess = str(f"True or False: {question}\n")
    send_to_all(client_sockets, mess) # f"True or False: {question}\n")
    curr_threads = []
    for idx, client_id in enumerate(client_sockets):
        sock = client_sockets[client_id]
        t = threading.Thread(target=play_trivia, args=(client_id, sock, question, client_answers))
        t.start()
        curr_threads.append(t)
    for t in curr_threads:
        t.join()
    if everyone_wrong_or_right(list(client_answers.values())):
        print('Nobody is eliminated this round! Let"s move to the next round.')
        send_to_all(client_sockets, 'Nobody is eliminated this round! Let"s move to the next round.')
        time.sleep(1)
        return True
    time.sleep(2)
    mess = ''
    if len(pid_to_name) > 2:
        for player in client_answers.keys():
            print(f'{pid_to_name[player]} {client_answers[player]}')
            mess += f'{pid_to_name[player]} {client_answers[player]}\n'
    else:
        for player in client_answers.keys():
            ans = client_answers[player]
            if ans == "is correct!":
                print(f'{pid_to_name[player]} is correct! {pid_to_name[player]} wins!')
                mess += f'{pid_to_name[player]} is correct! {pid_to_name[player]} wins!\n'
            else:
                print(f'{pid_to_name[player]} {client_answers[player]}')
                mess += f'{pid_to_name[player]} {client_answers[player]}\n'

    send_to_all(client_sockets, mess)
    time.sleep(1)

    for player in client_answers.keys():
        if client_answers[player] != 'is correct!':
            t = threading.Thread(target=elimination_msg, args=(client_sockets[player], pid_to_name[player]))
            t.start()
            client_sockets.pop(player)
            pid_to_name.pop(player)
            t.join()


def check_if_disconnected(client_sockets, client_sockets_og, players):
    """
    Checks if any clients have disconnected from the server.
    :param client_sockets: dict of client sockets still active in the game
    :param client_sockets_og: dict of client socket, original client socket pairs
    :param players: players in the game
    :return:
    """
    for client_id, sock in client_sockets.copy().items():
        try:
            # Check if the socket is still connected
            sock.sendall("".encode())
        except OSError:
            # If an OSError is raised, the socket is no longer connected
            print(f"{players[client_id]} has disconnected.")
            client_sockets.pop(client_id)
            client_sockets_og.pop(client_id)
            players.pop(client_id)


def everyone_wrong_or_right(lst):
    """
    Checks if everyone in the list is either all correct or all wrong.
    :param lst: list of client answers
    :return: True if everyone is either all correct or all wrong, False otherwise
    """
    return all([x != 'is correct!' for x in lst]) or all([x == 'is correct!' for x in lst])


def safe_sendall(client_id, sock, message):
    """
    Safely sends a message to a client.
    :param client_id: client id
    :param sock: socket to send message to
    :param message: message to send
    :return:
    """
    try:
        sock.sendall(message.encode())
    except (OSError, socket.timeout, ConnectionResetError, ConnectionAbortedError, BrokenPipeError) as e:
        if client_id in pid_to_name:
            global disconnected_clients
            disconnected_clients[client_id] = pid_to_name[client_id]


def send_to_all(list_of_sockets, msg):
    """
    Sends a message to all clients.
    :param list_of_sockets: list of sockets to send a message to
    :param msg: message to send
    :return:
    """
    try:
        curr_threads = []
        curr = None
        for client_id, sock in list_of_sockets.copy().items():
            curr = client_id
            t = threading.Thread(target=lambda sockt, message: safe_sendall(client_id, sockt, msg), args=(sock, msg))
            t.start()
            curr_threads.append(t)
        for t in curr_threads:
            t.join()
    except (OSError, socket.timeout, ConnectionResetError, ConnectionAbortedError, BrokenPipeError) as e:
        if curr in pid_to_name:
            global disconnected_clients
            disconnected_clients[curr] = pid_to_name[curr]


def elimination_msg(conn, name):
    """
    Sends an elimination message to a client.
    :param conn: socket connection
    :param name: name of the client
    :return:
    """
    try:
        conn.sendall(f"Sorry {name}, you are out of the game!\nPlease wait for the final results.\n".encode())
    except (OSError, socket.timeout, ConnectionResetError, ConnectionAbortedError, BrokenPipeError):
        pass


def send_question(client_id, conn, question):
    """
    Sends a question to a client.
    :param client_id: client id
    :param conn: socket connection
    :param question: question to send
    :return:
    """
    try:
        conn.sendall(f'True or False: {question}\n'.encode())
    except (OSError, socket.timeout, ConnectionResetError, ConnectionAbortedError, BrokenPipeError):
        if client_id in pid_to_name:
            global disconnected_clients
            disconnected_clients[client_id] = pid_to_name[client_id]


def play_trivia(client_id, conn, question, client_answers):
    """
    Plays the trivia game with a client. Receives the client's answer and checks if it is correct.
    :param client_id: client id
    :param conn: socket connection
    :param question: question in the trivia game
    :param client_answers: dictionary of client answers
    :return:
    """
    global disconnected_clients
    try:
        conn.settimeout(15)
        ans = conn.recv(1024).decode().strip().upper()
        if ans == "":
            disconnected_clients[client_id] = pid_to_name[client_id]
            return
        client_answer = acceptable_answers.get(ans, 'bad answer')
        correct_answer = 'T' if questions[question] else 'F'

        if client_answer is None:
            client_answers[client_id] = 'did not answer in time !'
        elif client_answer == 'bad answer':
            client_answers[client_id] = 'gave an invalid input !'
        elif client_answer == correct_answer:
            client_answers[client_id] = 'is correct!'
        else:
            client_answers[client_id] = 'is incorrect!'
    except (socket.timeout, ConnectionResetError, ConnectionAbortedError, BrokenPipeError):
        # client_answers[client_id] = 'incorrect'
        if client_id in pid_to_name:
            disconnected_clients[client_id] = pid_to_name[client_id]


if __name__ == "__main__":
    # Start the server
    try:
        thread_a = threading.Thread(target=udp_broadcast)
        thread_b = threading.Thread(target=tcp_server, args=(LOCAL_IP, available_port, False, 2))
        # Start both threads
        thread_a.start()
        thread_b.start()

        thread_a.join()
        thread_b.join()
    except KeyboardInterrupt:
        pass
    finally:
        print("Server is shutting down...Goodbye!")
